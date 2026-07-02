import openpyxl
import re

wb = openpyxl.load_workbook(r'D:\data scraping ian.xlsx')
ws = wb['Sheet1']
all_lines = []
for i in range(1, ws.max_row + 1):
    v = ws.cell(row=i, column=1).value
    if v is not None:
        if isinstance(v, str):
            all_lines.append(v)
        else:
            all_lines.append(str(v))

text = '\n'.join(all_lines)

# Find I and N pairs
pattern = r'I: (\d+),\n\s*N: (\d+),'
matches = re.findall(pattern, text)
unique_n = set()
for i, n in matches:
    unique_n.add(int(n))
print(f'Sample N (sport/league IDs): {sorted(unique_n)[:30]}')

# Find PN values
pn_matches = re.findall(r'PN: "([^"]*)"', text)
print(f'\nPN values (period names): {set(pn_matches)}')

# Find TG values
tg_matches = re.findall(r'TG: "([^"]*)"', text)
print(f'\nTG values (tournament groups): {set(tg_matches)}')

# Find top-level structure
print(f'\nTotal lines: {len(all_lines)}')
print(f'Total chars: {len(text)}')

# Print last few lines
print('\nLast 10 lines:')
for line in all_lines[-10:]:
    print(f'  {line}')

# Check what N=214429 looks like (football?)
print('\n--- Items with N=214429 (likely football) ---')
for line in all_lines:
    if 'N: 214429' in line:
        print(line)
