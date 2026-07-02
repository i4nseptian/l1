import openpyxl, re

wb = openpyxl.load_workbook(r'D:\data scraping ian.xlsx')
ws = wb['Sheet1']
print(f'Total rows: {ws.max_row}')

lines = []
for i in range(1, ws.max_row + 1):
    v = ws.cell(row=i, column=1).value
    if v is not None:
        lines.append(str(v))

# Count SG entries
sports = set()
for line in lines:
    m = re.match(r'^\s*N:\s*(\d+)', line)
    if m:
        sports.add(m.group(1))
print(f'\nUnique sport IDs (N): {len(sports)}')

# Count TG entries
tgs = set()
for line in lines:
    m = re.match(r'^\s*TG:\s*"([^"]+)"', line)
    if m:
        tgs.add(m.group(1))
print(f'Unique TGs: {len(tgs)}')

# Count total MEC entries (markets count)
mec_count = 0
in_mec = False
for line in lines:
    if re.match(r'^\s*MEC:\s*\[', line):
        in_mec = True
        continue
    if in_mec:
        if re.match(r'^\s*\]', line):
            in_mec = False
            continue
        if '{' in line:
            mec_count += 1
print(f'Total MEC blocks: {mec_count}')

# Count CI entries
cis = set()
for line in lines:
    m = re.match(r'^\s*CI:\s*(\d+)', line)
    if m:
        cis.add(m.group(1))
print(f'Unique CIs (championships): {len(cis)}')

# Check total EC sum from SG level
ec_total = 0
for line in lines:
    m = re.match(r'^\s*EC:\s*(\d+)', line)
    if m:
        ec_total += int(m.group(1))
print(f'Total EC (event count): {ec_total}')

# Check I entries (IDs)
ids = []
for line in lines:
    m = re.match(r'^\s*I:\s*(\d+)', line)
    if m:
        ids.append(int(m.group(1)))
print(f'Total I entries (IDs): {len(ids)}')
print(f'ID range: {min(ids)} - {max(ids)}')
