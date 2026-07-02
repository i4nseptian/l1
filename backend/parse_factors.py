import openpyxl
import json
import re
import os

def parse_excel_to_json(excel_path, output_path):
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Sheet1']

    all_lines = []
    for i in range(1, ws.max_row + 1):
        v = ws.cell(row=i, column=1).value
        if v is not None:
            all_lines.append(str(v))

    text = '\n'.join(all_lines)

    # Quote unquoted keys at line start
    text = re.sub(r'(?m)^\s*([a-zA-Z_]\w*)\s*:', r'"\1":', text)

    # Quote the 3 specific unquoted values found in the data
    text = re.sub(r'(?m)^(5af2b330494765f3ca3d26bc)\s*,?\s*$', r'"\1",', text)
    text = re.sub(r'(?m)^(12763\.png)\s*,?\s*$', r'"\1",', text)
    text = re.sub(r'(?m)^(13249\.png)\s*,?\s*$', r'"\1",', text)

    # Normalize booleans
    text = re.sub(r':\s*\btrue\b', ': true', text)
    text = re.sub(r':\s*\bfalse\b', ': false', text)
    text = re.sub(r':\s*\bnull\b', ': null', text)

    # Remove trailing commas before closing brackets
    text = re.sub(r',\s*(\})', r'\1', text)
    text = re.sub(r',\s*(\])', r'\1', text)

    data = json.loads(text)

    sg_list = data.get('Value', {}).get('SG', [])

    tgs = set()
    markets_summary = []
    for item in sg_list:
        tg = item.get('TG', '') or item.get('PN', '') or 'Unknown'
        tgs.add(tg)
        for mec in item.get('MEC', []):
            markets_summary.append({
                'sport_id': item.get('N'),
                'period': item.get('P'),
                'period_name': item.get('PN', ''),
                'group': tg,
                'market_type': mec.get('MT'),
                'market_name': mec.get('N', ''),
                'event_count': mec.get('EC', 0)
            })

    result = {
        'total_sports': len(sg_list),
        'total_markets': len(markets_summary),
        'tournament_groups': sorted(tgs),
        'markets': markets_summary
    }

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"Parsed {len(sg_list)} sports, {len(markets_summary)} markets")
    print(f"Tournament groups: {sorted(tgs)}")
    return result

if __name__ == '__main__':
    excel_path = r'D:\data scraping ian.xlsx'
    output_path = r'D:\scraping\backend\data\factors.json'
    parse_excel_to_json(excel_path, output_path)
